from openpyxl import load_workbook
import pandas as pd
import streamlit as st


def get_bottom_right_position(df):
    """Returns the 1-based (row, column) position of the bottom-right cell in a DataFrame."""
    if df.empty:
        return None  # Handle empty DataFrame
    return [len(df), len(df.columns)] # 1-based index


def transform_standard_to_graph(df):
    node_df = [["guid", "name", "desc"]]
    rel_df = [["guid", "from", "to", "type"]]
    rt_df = [["guid", "from", "desc"]]
    for record in df:
        print(record)
    # funktioniert noch nicht.
    return [node_df, rel_df, rt_df]

def transform_2d_to_standard(df, data_start=[0, 1], data_end=[]):
    """
    Convert a 2D table with named columns into a standardized long format.

    Parameters:
        df (pd.DataFrame): Input DataFrame with names as row indices and projects as columns.
        data_start: the first cell to contain DATA VALUES, not metadata
        data_end, the last cell fo the table
    Returns:
        pd.DataFrame: Transformed DataFrame with columns like ['rowid', 'data', 'Name', 'Projekt'].

    """
    if data_end == []:
        data_end = get_bottom_right_position(df)

    df_long = [["og_cell", "value", "X-Dimension", "Y-Dimension"]]
    for i in range(data_start[0], data_end[0]):
        for j in range(data_start[1], data_end[1]):
            row = [f"{i}_{j}", df.iloc[i, j], df.iloc[data_start[0]-1, j], df.iloc[i, data_start[1]-1]]
            df_long.append(row)
    return pd.DataFrame(df_long)
def highlight_headers(df, header_locs, color="yellow"):
    """Highlight the detected headers based on their location."""
    def highlight(row):
        return [
            f"background-color: {color};" if (row.name, col_idx) in header_locs else ""
            for col_idx in range(len(df.columns))
        ]
    return df.style.apply(highlight, axis=1)
def get_headers(uploaded_file):
    if uploaded_file:
        # Read the Excel file and get sheet names
        excel_data = pd.ExcelFile(uploaded_file)
        sheet_names = excel_data.sheet_names

        detected_headers = {}
        detected_header_vals = {}

        # Create tabs for each sheet
        tabs = st.tabs(sheet_names)

        for i, sheet_name in enumerate(sheet_names):
            with tabs[i]:
                st.subheader(f"Sheet: {sheet_name}")

                # Read the sheet as a DataFrame
                sheet_df = pd.read_excel(excel_data, sheet_name=sheet_name, header=None)

                # Allow user to select the header row
                max_rows = min(len(sheet_df), 10)
                header_row = st.selectbox(
                    f"Select the header row for {sheet_name}",
                    options=list(range(max_rows)),
                    index=0,
                    key=f"header_row_{sheet_name}"
                )

                # Allow user to select if there is a header column (2D table)
                headers_in_row = st.toggle(f"Does {sheet_name} contain headers in a row?", key=f"headers_in_row_{sheet_name}", value=True)
                headers_in_col = st.toggle(f"Does {sheet_name} contain headers in a column?", key=f"headers_in_col_{sheet_name}")

                # Extract header locations based on user input
                header_locs = []
                header_vals = []
                if headers_in_col:
                    # Allow user to select the header row
                    max_cols = min(len(sheet_df), 10)
                    header_col = st.selectbox(
                        f"Select the header col for {sheet_name}",
                        options=list(range(max_cols)),
                        index=0,
                        key=f"header_col_{sheet_name}"
                    )
                    for row in range(len(list(sheet_df.iterrows()))):
                        header_locs.append((row, header_col))

                if headers_in_row:
                    for col in range(len(sheet_df.columns)):
                        header_locs.append((header_row, col))
                        header_vals.append(sheet_df.iloc[header_row, col])

                detected_headers[sheet_name] = header_locs
                detected_header_vals[sheet_name] = header_vals

                if headers_in_col:
                    highlighted_df = highlight_headers(sheet_df, header_locs, color="orange")
                    st.write("Uploaded DataFrame:")
                    st.dataframe(highlighted_df)

                    sheet_df = transform_2d_to_standard(sheet_df, data_start=[header_col+1, header_row+1])
                    detected_headers[sheet_name] = [(0, 0), (0, 1), (0, 2), (0, 3)]
                    detected_header_vals[sheet_name] = ["og_cell", "value", "X-Dimension", "Y-Dimension"]


                # Highlight headers and display the DataFrame
                highlighted_df = highlight_headers(sheet_df, detected_headers[sheet_name])
                st.write("transformed DataFrame:")
                st.dataframe(highlighted_df)

        # Button to confirm headers
        if st.button("Confirm Headers for all sheets"):
            return detected_header_vals

# Function to extract entity names and attribute names
def extract_entity_attributes(file, structure="other"):
    entity_details = {}

    if file.name.endswith('.csv'):
        # Read CSV and use the filename as the entity name
        df = pd.read_csv(file)
        entity_name = file.name.split('.')[0]
        attributes = df.columns.tolist()
        entity_details[entity_name] = attributes

    elif file.name.endswith(('.xls', '.xlsx')):
        # Load Excel file
        workbook = load_workbook(file, data_only=True)
        for sheet in workbook.sheetnames:
            sheet_data = pd.DataFrame(workbook[sheet].values)

            if sheet_data.empty:
                continue

            headers = False
            if structure == "other":
                # Check if there's a clear header (row 1)
                headers = sheet_data.iloc[0].dropna().tolist()
            elif structure == "row0":
                headers = sheet_data.iloc[0].dropna().tolist()
            elif structure == "row1":
                headers = sheet_data.iloc[1].dropna().tolist()
            if headers:
                entity_details[sheet] = headers
            else:
                # Fallback for undefined headers (use indexes)
                headers = [f"Column {i + 1}" for i in range(sheet_data.shape[1])]
                entity_details[sheet] = headers

    else:
        st.warning("Uploaded file format is not supported. Please upload a CSV or Excel file.")

    return entity_details
def execute_gto_transformation(data, mapper, gto):
    pass
def execute_ziel_transformation(data, mapper, ziel):
    pass
def execute_mapper_transformation(data, mapper):
    unique_quell_sheets = []
    for n in range(len(mapper)):
        sheetname = mapper[n]["Quelle_Sheet"]
        if sheetname not in unique_quell_sheets:
            unique_quell_sheets.append(sheetname)
    quelle_df = pd.read_excel(data, sheet_name=unique_quell_sheets)

    ziel_df = pd.DataFrame()

    for entry in mapper:
        try:
            quelle_sheet = entry['Quelle_Sheet']
            quelle_column = entry['Quelle_Column']
            transformation_rule = entry['Transformation_Rule']
            transformation_param = entry['Transformation_Rule_param']
            ziel_sheet = entry['Ziel_Sheet']
            ziel_column = entry['Ziel_Column']

            # Get the relevant source DataFrame
            source_df = quelle_df[quelle_sheet]

            # Perform the transformation based on the rule
            if transformation_rule == '1:1':
                ziel_df[ziel_column] = source_df[quelle_column]
            elif transformation_rule == 'multiply':
                ziel_df[ziel_column] = source_df[quelle_column] * float(transformation_param)
            elif transformation_rule == "cut_left":
                ziel_df[ziel_column] = source_df[quelle_column].str.split(transformation_param, expand=True)
            elif transformation_rule == "cut_sep_right":
                # LINKS RECHTS LOGIK FEHLT
                ziel_df[ziel_column] = source_df[quelle_column].str.split(transformation_param, expand=True)
            elif transformation_rule == "cut_sep_left":
                ziel_df[ziel_column] = source_df[quelle_column].str.split(transformation_param, expand=True)
            elif transformation_rule == "cut_right":
                ziel_df[ziel_column] = source_df[quelle_column].str.split(transformation_param, expand=True)
            elif transformation_rule == "concat":
                order = [int(i) for i in transformation_param.split(",")]
                ziel_df[ziel_column] = source_df.iloc[:, order].apply(lambda x: "".join(x.astype(str)), axis=1)
            elif transformation_rule == "add":
                ziel_df[ziel_column] = source_df[quelle_column] + float(transformation_param)
        except Exception as e:
            return [False, e]
    return [True, ziel_df]


def execute_transformation_old(data, mapper, ziel):
    ziel_preview = []
    ziel_df = pd.DataFrame(columns=ziel)



    unique_quell_sheets = []
    for n in range(len(mapper)):
        sheetname = mapper[n]["Quelle_Sheet"]
        if sheetname not in unique_quell_sheets:
            unique_quell_sheets.append(sheetname)

    quelle_df = pd.read_excel(data, sheet_name=unique_quell_sheets)
    for entity in quelle_df:
        for record in entity:
            st.write(record)
    """
        quelle = pd.read_excel(data, sheet_name=mapper[row]["Quelle_Sheet"])
        #quelle_df = pd.read_excel(data, sheet_name=mapper)

    st.write(ziel_df)

    mapping = []
    quelle_column = mapping["Quelle_Column"]
    ziel_column = mapping["Ziel_Column"]
    rule = mapping["Transformation_Rule"]
    rule_param = mapping["Transformation_Rule_param"]

    if rule == "1:1":
        ziel_df[ziel_column] = quelle_df[quelle_column]

    ziel_preview.append(ziel_df)

    # Combine and display the preview of ZIEL
    combined_preview = pd.concat(ziel_preview)
    st.subheader("ZIEL Preview")
    st.dataframe(combined_preview)

    # Provide a download option for the transformed ZIEL
    ziel_csv = combined_preview.to_csv(index=False)
    st.download_button(
        "Download Transformed ZIEL",
        ziel_csv,
        "transformed_ziel.csv",
        "text/csv",
        key="download_ziel_csv"
    )



# Streamlit app
def main():
    st.title("Entity and Attribute Extractor")

    uploaded_file = st.file_uploader("Upload a CSV or Excel file", type=["csv", "xls", "xlsx"])

    if uploaded_file:
        with st.spinner("Processing file..."):
            entity_details = extract_entity_attributes(uploaded_file)

        if entity_details:
            st.success("Entities and attributes extracted successfully!")
            for entity, attributes in entity_details.items():
                st.subheader(f"Entity: {entity}")
                st.write("Attributes:")
                st.write(attributes)
        else:
            st.warning("No entities or attributes found in the uploaded file.")


if __name__ == "__main__":
    #main()
    my_mapper = [{'Quelle_Sheet': 'meine_räume', 'Quelle_Column': None, 'Transformation_Rule': None,
                  'Transformation_Rule_param': None, 'Ziel_Sheet': 'ROOMS', 'Ziel_Column': None},
                 {'Quelle_Sheet': 'meine_räume', 'Quelle_Column': 'Fläche', 'Transformation_Rule':
                     'multiply', 'Transformation_Rule_param': '100', 'Ziel_Sheet': 'ROOMS', 'Ziel_Column': 'Bodenfläche'}]
    print(len(my_mapper))



def other_rules():
   


"""


if __name__ == '__main__':
    pass