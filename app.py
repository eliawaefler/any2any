"""
app.py ist das forntend 
"""

from openpyxl import Workbook


def create_excel_with_cell_names():
    # Create a new workbook
    wb = Workbook()

    # Add a sheet named "testsheet"
    ws = wb.active
    ws.title = "testsheet"

    # Fill the first 5x5 cells with cell names
    for row in range(1, 6):  # Rows 1 through 5
        for col in range(1, 6):  # Columns 1 through 5
            cell_name = ws.cell(row=row, column=col).coordinate
            ws.cell(row=row, column=col).value = cell_name

    # Save the workbook to text.xlsx
    wb.save("text.xlsx")
    print("File 'text.xlsx' with sheet 'testsheet' created successfully.")


# Call the function to create the file
create_excel_with_cell_names()
