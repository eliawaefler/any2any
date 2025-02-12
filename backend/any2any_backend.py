
"""
hier kommen funktionen, die von der ANWENDUNG aufgerufen werden.
"""
import numpy as np
import pandas
from openpyxl import load_workbook
import pandas as pd


def get_bottom_right_position(df):
    """Returns the 1-based (row, column) position of the bottom-right cell in a DataFrame."""
    if df.empty:
        return None  # Handle empty DataFrame
    return [len(df), len(df.columns)] # 1-based index


def transform_standard_to_graph(df):
    node_df = [["guid", "name", "desc"]]
    rel_df = [["guid", "from", "to", "type"]]
    rt_df = [["guid", "from", "desc"]]
    for record in df:
        print(record)
    # funktioniert noch nicht.
    return [node_df, rel_df, rt_df]


def get_cells_in_range(start, end):
    """ Returns a list of cell coordinates within a given range. """
    cells = []
    for x in range(int(start[0]), int(end[0] + 1)):
        for y in range(int(start[1]), int(end[1] + 1)):
            cells.append(tuple([x, y]))
    return cells


def transform_2d_to_standard(df, header_names, data_start):
    """ Transforms a multi-dimensional data format into a standard 2D DataFrame. """
    print(df)
    #pandas.DataFrame(columns=header_names, data=df.iloc[data_start[0]:, data_start[1]:].copy())
    transformed_df = pandas.DataFrame(data=df.iloc[data_start[0]:, data_start[1]:].copy()).melt()
    #transformed_df = df.melt()
    print(transformed_df)

    transformed_df.columns = header_names
    return transformed_df


def old_highlight_multiple_cells(df, highlights):
    """ Highlights specific cells in a DataFrame. """
    styled_df = df.style.applymap(lambda _: "background-color: yellow", subset=highlights)
    return styled_df


def highlight_multiple_cells(df, highlights, highlight_headers=False):
    """
    Highlights specific cells and optionally headers in a DataFrame.

    Parameters:
        df (pd.DataFrame): The DataFrame to style.
        highlights (list of tuples): List of (row, col) tuples to highlight.
        highlight_headers (bool): Whether to highlight column headers.

    Returns:
        pd.io.formats.style.Styler: Styled DataFrame.
    """

    def highlight_func(val):
        return "background-color: yellow" if val else ""

    # Create a DataFrame of same shape as df, filled with empty strings
    highlight_df = pd.DataFrame("", index=df.index, columns=df.columns)

    # Apply highlights to specified cells
    for row, col in highlights:
        if row in df.index and col in df.columns:
            highlight_df.loc[row, col] = "background-color: yellow"

    styled_df = df.style.apply(lambda _: highlight_df, axis=None)

    # If headers need to be highlighted
    if highlight_headers:
        styled_df.set_table_styles([
            {"selector": f"th.col{df.columns.get_loc(col)}", "props": "background-color: yellow;"}
            for _, col in highlights if col in df.columns
        ], overwrite=False)

    return styled_df


# Function to extract entity names and attribute names
def extract_entity_attributes(file, structure="other"):
    entity_details = {}
    if file.name.endswith('.csv'):
        # Read CSV and use the filename as the entity name
        df = pd.read_csv(file)
        entity_name = file.name.split('.')[0]
        attributes = df.columns.tolist()
        entity_details[entity_name] = attributes

    elif file.name.endswith(('.xls', '.xlsx')):
        # Load Excel file
        workbook = load_workbook(file, data_only=True)
        for sheet in workbook.sheetnames:
            sheet_data = pd.DataFrame(workbook[sheet].values)

            if sheet_data.empty:
                continue

            headers = False
            if structure == "other":
                # Check if there's a clear header (row 1)
                headers = sheet_data.iloc[0].dropna().tolist()
            elif structure == "row0":
                headers = sheet_data.iloc[0].dropna().tolist()
            elif structure == "row1":
                headers = sheet_data.iloc[1].dropna().tolist()
            if headers:
                entity_details[sheet] = headers
            else:
                # Fallback for undefined headers (use indexes)
                headers = [f"Column {i + 1}" for i in range(sheet_data.shape[1])]
                entity_details[sheet] = headers

    else:
        raise TypeError
        # return "Uploaded file format is not supported. Please upload a CSV or Excel file."

    return entity_details


def execute_gto_transformation(data, mapper, gto):
    if data:
        if mapper:
            if gto:
                pass
    pass


def execute_ziel_transformation(data, mapper, ziel):
    if data:
        if mapper:
            if ziel:
                pass
    pass


def execute_mapper_transformation(data, mapper):

    """
    :param data:
    :param mapper:
    :return:
    """

    unique_quell_sheets = []
    for n in range(len(mapper)):
        sheet_name = mapper[n]["Quelle_Sheet"]
        if sheet_name not in unique_quell_sheets:
            unique_quell_sheets.append(sheet_name)
    quelle_df = pd.read_excel(data, sheet_name=unique_quell_sheets)

    ziel_df = pd.DataFrame()

    for entry in mapper:
        try:
            quelle_sheet = entry['Quelle_Sheet']
            quelle_column = entry['Quelle_Column']
            transformation_rule = entry['Transformation_Rule']
            transformation_param = entry['Transformation_Rule_param']
            ziel_sheet = entry['Ziel_Sheet']
            ziel_column = entry['Ziel_Column']

            # Get the relevant source DataFrame
            source_df = quelle_df[quelle_sheet]

            # Perform the transformation based on the rule
            if transformation_rule == '1:1':
                ziel_df[ziel_column] = source_df[quelle_column]
            elif transformation_rule == 'multiply':
                ziel_df[ziel_column] = source_df[quelle_column] * float(transformation_param)
            elif transformation_rule == "cut_left":
                ziel_df[ziel_column] = source_df[quelle_column].str.split(transformation_param, expand=True)
            elif transformation_rule == "cut_sep_right":
                # LINKS RECHTS LOGIK FEHLT
                ziel_df[ziel_column] = source_df[quelle_column].str.split(transformation_param, expand=True)
            elif transformation_rule == "cut_sep_left":
                ziel_df[ziel_column] = source_df[quelle_column].str.split(transformation_param, expand=True)
            elif transformation_rule == "cut_right":
                ziel_df[ziel_column] = source_df[quelle_column].str.split(transformation_param, expand=True)
            elif transformation_rule == "concat":
                order = [int(i) for i in transformation_param.split(",")]
                ziel_df[ziel_column] = source_df.iloc[:, order].apply(lambda x: "".join(x.astype(str)), axis=1)
            elif transformation_rule == "add":
                ziel_df[ziel_column] = source_df[quelle_column] + float(transformation_param)
        except Exception as e:
            return [False, e]
    return [True, ziel_df]







def transform_excel_data(file_path: str, sheet_name: str):
    """
    Reads an Excel sheet and transforms the data such that for every row,
    if a value exists in columns C, D, ..., it creates a new row where:
    - Column A is copied.
    - The value from C, D, etc., is moved to column B.

    :param file_path: Path to the Excel file.
    :param sheet_name: Name of the sheet to read.
    :return: Transformed Pandas DataFrame.
    """
    # Read the Excel file
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    print(df)
    # Ensure we have at least two columns
    if df.shape[1] < 2:
        raise ValueError("The sheet must have at least two columns (A and B).")

    # Get column names
    col_a, col_b = df.columns[0], df.columns[1]
    other_cols = df.columns[2:]  # All columns except A and B

    transformed_data = []

    for _, row in df.iterrows():
        a_value = row[col_a]
        for col in other_cols:
            if pd.notna(row[col]):  # Check if the cell is not empty
                transformed_data.append({col_a: a_value, col_b: row[col]})

    # Create a new DataFrame
    new_df = pd.DataFrame(transformed_data)

    return new_df


# Function to extract sheet names and column headers from an uploaded Excel file
def extract_file_structure(uploaded_file):
    structure = {}
    excel_file = pd.ExcelFile(uploaded_file)
    for sheet in excel_file.sheet_names:
        df = excel_file.parse(sheet)
        structure[sheet] = list(df.columns)
    return structure



def standard_tabelle():


    # A     B       C       D       E           F               G           H           ...
    # ID    GUID    name    wert    Einheit     dim1            dim1        dim3        ...
    # 01    A1B2C3  p1      5       h           G2             19.01.2025  projekt A   ...

    # Zeiterfassung
    # 01, A1B2C3, Elia, 5, h, Zeiterfassung, 09.01.2025, Projekt1,  ...

    # Energiedashboard
    # 01, A1B2C3, 5, kWh, Strom, 01.01.2025

    # tabellenname

    pass



def finde_tabelle():
    pass


def standardize_dataframe(df, metadata_dict):
    data_records = []

    # Create a metadata lookup dictionary
    metadata_lookup = {key: {} for key in metadata_dict}

    # Populate metadata lookup
    for key, locations in metadata_dict.items():
        for row, col in locations:
            loc_key = f"{row}_{col}"
            metadata_lookup[key][loc_key] = df.iat[row, col]

    # Process all cells that are **not** metadata
    for row in range(df.shape[0]):
        for col in range(df.shape[1]):
            loc_key = f"{row}_{col}"

            # Skip metadata locations
            if any(loc_key in meta_dict for meta_dict in metadata_lookup.values()):
                continue

            # Build row data
            record = {
                "og_loc": loc_key,
                "val": df.iat[row, col]
            }

            # Attach metadata from any matching locations
            for key in metadata_lookup:
                # Check if the metadata exists for this row/col
                metadata_value = None
                for meta_loc, meta_value in metadata_lookup[key].items():
                    meta_row, meta_col = map(int, meta_loc.split("_"))

                    # If metadata is in the **same column**, apply it
                    if meta_col == col:
                        metadata_value = meta_value
                    # If metadata is in the **same row**, apply it
                    elif meta_row == row:
                        metadata_value = meta_value

                record[key] = metadata_value
            if record["val"] not in [None, np.nan]:
                data_records.append(record)

    # Convert list to DataFrame
    standardized_df = pd.DataFrame(data_records).dropna(subset=["val"])


    # Ensure correct column order: Location, Value, then metadata fields
    final_columns = ["og_loc", "val"] + list(metadata_dict.keys())
    standardized_df = standardized_df[final_columns]

    return standardized_df


if __name__ == '__main__':
    # Example Usage
    df = pd.DataFrame([
        ["X1", "X2", "X3", None],  # Column Metadata
        [10, 20, 30, "A"],  # Data + Row Metadata
        [40, 50, 60, "B"]  # Data + Row Metadata
    ])

    metadata_dict = {
        "Category": [(0, 0), (0, 1), (0, 2)],  # Column metadata
        "Group": [(1, 3), (2, 3)]  # Row metadata
    }

    standardized_df = standardize_dataframe(df, metadata_dict)
    print(standardized_df)
